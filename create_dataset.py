from openpyxl import load_workbook
import pandas as pd


def get_dataset(
    excel, 
    col_text="Question", 
    col_entity="text_with_entity",
    min_example_num=5,
    df_dataset=None
    ):

    # Load the Excel file
    workbook = load_workbook(excel)

    # Get the sheet names
    sheet_names = workbook.sheetnames

    try:
        if not df_dataset:
            df_dataset = pd.Series()
    except:
        pass

    for sheet_name in sheet_names:

        if sheet_name.lower() == "summary":
            continue
        
        # get intent from sheet name
        intent = sheet_name.split(".")[-1]
        intent = intent.replace("-", "")
        intent = intent.replace(",", "")
        intent = "_".join(intent.split())
        intent = intent.lower()
        
        try:
            df_temp = pd.read_excel(excel, sheet_name=sheet_name)
            df_temp = df_temp[df_temp[col_text].notnull() & df_temp[col_entity].notnull()]

            # get text
            texts = df_temp[col_text].tolist()
            texts = [text.replace("\n", "").strip().lower() for text in texts]
            
            # get entity
            entities = df_temp[col_entity].tolist()
            entities = [entity.replace("\n", "").strip().lower() for entity in entities]
            
            # create text column
            text_col = pd.Series(texts)

            # create entity column
            entity_col = pd.Series(entities)
            
            # create label column
            label_col = pd.Series(intent for _ in range(len(texts)))

            if len(text_col)>=min_example_num:
                # create dataframe
                df_dataset_temp = pd.DataFrame({"text": text_col, "text_with_entity": entity_col, "label": label_col})
                df_dataset = pd.concat([df_dataset, df_dataset_temp]).reset_index(drop=True)
            
        except Exception as e:
            print(excel, sheet_name)
            print(e)
            pass
        
    
    df_dataset = df_dataset.reset_index(drop=True)

    return df_dataset


if __name__ == "__main__":
    df = get_dataset("data/from_ct/20231213/IntentChinese.xlsx")
    df = get_dataset("data/from_ct/20231213/IntentSimplifiedChinese.xlsx", df_dataset=df)
    df = get_dataset("data/from_ct/20231213/IntentEnglish.xlsx", df_dataset=df)
    
    df.to_excel("data/from_ct/20231213/dataset.xlsx")

